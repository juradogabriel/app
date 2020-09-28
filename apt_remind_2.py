import re
import pandas as pd
import openpyxl

from flask import Flask, request
from twilio.twiml.voice_response import Gather, VoiceResponse
from twilio.twiml.messaging_response import Message, MessagingResponse
from twilio.rest import Client

from config import PRIVATE_NUMBER,TWILIO_NUMBER
from config import account_sid, auth_token

#import excel library
from openpyxl import load_workbook
from tempfile import NamedTemporaryFile

#Calls flaskapp, it listens to port 5000
app=Flask(__name__)

#summons flaskapp with/sms tag. Loops the sms function whenever 5K is initiated
@app.route('/sms',methods=['POST'])
def sms():
    from_number=request.form['From']
    msg_body=request.form['Body']

    print(from_number,' ',msg_body)

    if msg_body == '1':
        excel_reply(from_number, msg_body)
        return send_reply("Thanks! See you at the office - Dr. Sanchez",from_number)

    else:
        excel_reply(from_number, msg_body)
        return send_reply("We understand that plans change. Thank you for letting us know!",from_number)

def excel_reply(number,response):
    wb = load_workbook('apt_remnd.xlsx')
    ws = wb.active
    row = 2
    while row <= 12:
        number = ws.cell(row, column=3)
        n = "+" + str(number.value)
        if n == number:
            if response == '1':
                ws.cell(row=2, column=5).value = "Confirmed"
            else:
                ws.cell(row=2, column=5).value = "Canceled"
            row = row+50
        row = row+1



def call_people():
    wb = load_workbook('apt_remnd.xlsx')
    ws = wb.active
    row = 2
    while row <= 12:
        time = ws.cell(row, column=1)
        date = ws.cell(row = 2, column = 9)
        name = ws.cell(row, column=2)
        number = ws.cell(row, column=3)
        print(name.value,number.value)
        row = row+1
        h = "Hello "+str(name.value)+"! This is a reminder that you have an appointment scheduled "+str(date.value)+" at "+str(time.value)+" . Reply '1' to confirm or '2' to cancel."
        n = "+"+str(number.value)
        send_msg(h,n)



#replies on port 5K to twilio while port is open
  
def send_reply(msg,number):
  response=MessagingResponse()
  response.message(msg,to=number,from_=TWILIO_NUMBER)
  return str(response)

#sends a fresh message using your client credentials
def send_msg(msg,number):
  client=Client(account_sid,auth_token)

  message=client.messages \
	  .create(
	     body=msg,
	     from_=TWILIO_NUMBER,
	     to=number
	  )

  print(message.sid)

#main function, appointment reminder
if __name__=='__main__':
    call_people()

    app.run()
